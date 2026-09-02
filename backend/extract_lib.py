# -*- coding: utf-8 -*-
# Shared xlsx-parsing logic used by both the local extract.py CLI and the
# Fly.io upload endpoint (app.py). Kept dependency-free (just openpyxl) so it
# can run in either place unchanged.
from datetime import datetime
import re
import openpyxl

# B2C/FIN23 exports often carry 40+ columns (call logs, UTM tags, health scores, etc.)
# but app.js's buildRawData() only ever reads these. Dropping the rest keeps the
# payload small.
FIN23_COLUMNS = {
    'currentrmname', 'current rm name', 'rm', 'curren rm',
    'createddate', 'created date',
    'laststatusdate', 'last status date',
    'leadinprocessdate', 'leadprocessdate', 'lead in process date',
    'converteddate', 'converted date',
    'ctm', 'created month',
    'lpm', 'leadprocessmonth',
    'cm', 'converted month', 'convertedmonth',
    'fmonth',
    'team',
    'clientname', 'client name',
    'landingpage', 'landing page',
    'platformname', 'platform name',
    'campaign name', 'categoryname', 'campaignname', 'category name',
    'userid', 'user id',
    'leadhead', 'lead head',
    'leadstatus', 'lead status',
    'firstrmname', 'first rm name',
    'convertedbyname', 'converted by name',
    'annualincome', 'annual income',
    'clientcategory', 'client category',
}


def cell_value(v):
    """Mirror what SheetJS (raw:true, no cellDates) hands to app.js:
    strings/numbers pass through, datetimes become ISO strings, blanks become ''."""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    return v


def header_text(h):
    return '' if h is None else str(h).strip()


def rows_from_header_row(all_rows, header_idx, column_allowlist=None):
    headers = [header_text(h) for h in all_rows[header_idx]]
    if column_allowlist is not None:
        headers = [h if h.lower() in column_allowlist else '' for h in headers]
    out = []
    for row in all_rows[header_idx + 1:]:
        obj = {}
        any_val = False
        for c, h in enumerate(headers):
            if not h:
                continue
            v = row[c] if c < len(row) else None
            v = cell_value(v)
            obj[h] = v
            if v != '':
                any_val = True
        if any_val:
            out.append(obj)
    return out


def find_sheet(wb, *names):
    lower_targets = [n.lower() for n in names]
    for name in wb.sheetnames:
        if name.lower() in lower_targets:
            return name
    return wb.sheetnames[0]


def load_sheet_rows(file_like, sheet_names=None, column_allowlist=None):
    """file_like: a path string OR a file-like object (e.g. an uploaded file stream)."""
    wb = openpyxl.load_workbook(file_like, read_only=True, data_only=True)
    sheet_name = find_sheet(wb, *sheet_names) if sheet_names else wb.sheetnames[0]
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return []
    return rows_from_header_row(all_rows, 0, column_allowlist=column_allowlist)


def load_revenue_input_rows(file_like):
    """Scan the first 5 rows for a header row containing 'clientname', 'rm',
    and 'total' (lowercased exact match) — same detection the old in-browser
    parseRevenueInput() used."""
    wb = openpyxl.load_workbook(file_like, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return []
    header_idx = 0
    for i in range(min(len(all_rows), 5)):
        lowered = [header_text(c).lower() for c in all_rows[i]]
        if 'clientname' in lowered and 'rm' in lowered and 'total' in lowered:
            header_idx = i
            break
    return rows_from_header_row(all_rows, header_idx)


# ---------- BD Accountability Tracker ----------
# Each BD rep's leads live in a sheet literally named "<Person> Q<N>" (e.g.
# "Himani Q1"). New quarters/reps are picked up automatically by this naming
# pattern alone — no code change needed when "Himani Q3" is added later.
# Every other sheet in the workbook (dashboards, logs, pivot tables) is
# ignored since it won't match.
BD_SHEET_PATTERN = re.compile(r'^(.+?)\s+Q(\d+)$', re.IGNORECASE)

# dateAssigned/email/gmeetJoined/currentStage are read from the tracker.
# RM/Team come from the matching B2C record instead (see
# annotateBDWithB2CMatch in app-state.js) since that's the CRM's live
# ownership, not what the tracker originally logged; Financial Plan Created
# isn't tracked at all. dateAssigned is also the Month fallback for leads
# that don't match a B2C record. currentStage is kept as the BD Team's own
# self-reported progress — shown separately from the B2C-derived Stage (see
# the Workpoint Status / BD Team Status toggle on the dashboard) since the
# two can disagree once a lead is reassigned after the BD rep logged it.
# Source headers are inconsistent across sheets (trailing spaces, casing) —
# matched case/space-insensitively.
BD_TRACKER_COLUMNS = {
    'dateAssigned': {'date assigned'},
    'email':        {'email'},
    'gmeetJoined':  {'gmeet joined?'},
    'currentStage': {'current stage'},
}

# "Current Stage" free-text values seen in the wild, collapsed onto one
# canonical label per real-world stage. Unrecognized non-blank values pass
# through upper-cased rather than being silently dropped.
BD_STAGE_MAP = {
    'lead assigned': 'LEAD ASSIGNED', 'assign': 'LEAD ASSIGNED',
    'in follow-up': 'IN FOLLOW-UP', 'follow up': 'IN FOLLOW-UP',
    'in process': 'IN PROCESS',
    'dropped': 'DROPPED',
    'onhold/dead': 'ON HOLD/DEAD',
    'converted': 'CONVERTED',
    'tax filling done': 'TAX FILING DONE',
}

def normalize_bd_stage(v):
    s = header_text(v).strip().lower()
    if not s:
        return ''
    return BD_STAGE_MAP.get(s, header_text(v).strip().upper())

# GMeet Joined? collapses onto a 3-state Yes/No/Pending.
BD_YNP_MAP = {
    'yes': 'Yes', 'joined': 'Yes',
    'no': 'No', 'not joined': 'No',
    'pending': 'Pending',
}

def normalize_bd_ynp(v):
    s = header_text(v).strip().lower()
    return BD_YNP_MAP.get(s, '')


def load_bd_tracker_rows(file_like):
    wb = openpyxl.load_workbook(file_like, read_only=True, data_only=True)
    out = []
    for sheet_name in wb.sheetnames:
        m = BD_SHEET_PATTERN.match(sheet_name.strip())
        if not m:
            continue
        person, quarter_num = m.group(1).strip(), m.group(2)
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        headers = [header_text(h) for h in all_rows[0]]
        col_for_idx = {}
        for i, h in enumerate(headers):
            key = h.strip().lower()
            for canon, variants in BD_TRACKER_COLUMNS.items():
                if key in variants:
                    col_for_idx[i] = canon
                    break
        for row in all_rows[1:]:
            obj = {}
            any_val = False
            for i, canon in col_for_idx.items():
                v = row[i] if i < len(row) else None
                if canon == 'gmeetJoined':
                    v = normalize_bd_ynp(v)
                elif canon == 'currentStage':
                    v = normalize_bd_stage(v)
                else:
                    v = cell_value(v)
                obj[canon] = v
                if v not in ('', None):
                    any_val = True
            if not any_val:
                continue
            obj['person'] = person
            obj['quarter'] = 'Q' + quarter_num
            obj['sourceSheet'] = sheet_name
            out.append(obj)
    wb.close()
    return out
