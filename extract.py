# -*- coding: utf-8 -*-
# Interactively asks for each source file, reads it with openpyxl, and writes a
# compact data.js that the webpage loads directly — no in-browser XLSX parsing.
#
# B2C (FIN23 Lead Management) is required. Revenue Input, B2B, FY2026, and Plan
# Approval are all optional — skip any of them and the dashboard just shows
# "not found" for that data source, same as today.
#
# This script does NOT touch Cost Per Campaign or EMPLOYEE_REF — those still live
# in snapshot.js / the in-app editors, untouched by this flow.
import json, os, time
from datetime import datetime
import openpyxl

t0 = time.time()

# B2C/FIN23 exports often carry 40+ columns (call logs, UTM tags, health scores, etc.)
# but app.js's buildRawData() only ever reads these. Dropping the rest keeps data.js
# well under GitHub's 100MB per-file push limit for a large lead file.
FIN23_COLUMNS = {
    'currentrmname', 'current rm name', 'rm', 'curren rm',
    'createddate', 'created date',
    'laststatusdate', 'last status date',
    'leadinprocessdate', 'leadprocessdate', 'lead in process date',
    'converteddate', 'converted date',
    'ctm', 'created month',
    'lpm', 'leadprocessmonth',
    'cm', 'converted month', 'convertedmonth',
    'fmonth',
    'team',
    'clientname', 'client name',
    'landingpage', 'landing page',
    'platformname', 'platform name',
    'campaign name', 'categoryname', 'campaignname', 'category name',
    'userid', 'user id',
    'leadhead', 'lead head',
    'leadstatus', 'lead status',
    'firstrmname', 'first rm name',
    'convertedbyname', 'converted by name',
    'annualincome', 'annual income',
    'clientcategory', 'client category',
}

def cell_value(v):
    """Mirror what SheetJS (raw:true, no cellDates) hands to app.js:
    strings/numbers pass through, datetimes become ISO strings, blanks become ''."""
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d %H:%M:%S')
    return v

def header_text(h):
    return '' if h is None else str(h).strip()

def rows_from_header_row(all_rows, header_idx, column_allowlist=None):
    headers = [header_text(h) for h in all_rows[header_idx]]
    if column_allowlist is not None:
        headers = [h if h.lower() in column_allowlist else '' for h in headers]
    out = []
    for row in all_rows[header_idx+1:]:
        obj = {}
        any_val = False
        for c, h in enumerate(headers):
            if not h:
                continue
            v = row[c] if c < len(row) else None
            v = cell_value(v)
            obj[h] = v
            if v != '':
                any_val = True
        if any_val:
            out.append(obj)
    return out

def find_sheet(wb, *names):
    lower_targets = [n.lower() for n in names]
    for name in wb.sheetnames:
        if name.lower() in lower_targets:
            return name
    return wb.sheetnames[0]

def load_sheet_rows(path, sheet_names=None, column_allowlist=None):
    print(f'  Reading {path}...', end=' ', flush=True)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_name = find_sheet(wb, *sheet_names) if sheet_names else wb.sheetnames[0]
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        print('0 rows (empty sheet)', flush=True)
        return []
    rows = rows_from_header_row(all_rows, 0, column_allowlist=column_allowlist)
    print(f'{len(rows)} rows from sheet "{sheet_name}"  ({time.time()-t0:.1f}s)', flush=True)
    return rows

def load_revenue_input_rows(path):
    """Scan the first 5 rows for a header row containing 'clientname', 'rm',
    and 'total' (lowercased exact match) — same detection the old in-browser
    parseRevenueInput() used."""
    print(f'  Reading {path}...', end=' ', flush=True)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        print('0 rows (empty sheet)', flush=True)
        return []
    header_idx = 0
    for i in range(min(len(all_rows), 5)):
        lowered = [header_text(c).lower() for c in all_rows[i]]
        if 'clientname' in lowered and 'rm' in lowered and 'total' in lowered:
            header_idx = i
            break
    rows = rows_from_header_row(all_rows, header_idx)
    print(f'{len(rows)} rows  ({time.time()-t0:.1f}s)', flush=True)
    return rows

def prompt_path(label, required):
    suffix = '' if required else ' (press Enter to skip)'
    while True:
        raw = input(f'{label}{suffix}: ').strip().strip('"')
        if not raw:
            if required:
                print('  This file is required — please provide a path.')
                continue
            return None
        if not os.path.isfile(raw):
            print(f'  File not found: {raw}')
            continue
        return raw

print('Marketing MIS — data.js builder')
print('Provide a path to each source file. Only B2C is required; skip the rest with Enter.')
print()

fin23_path = prompt_path('B2C — FIN23 Lead Management [REQUIRED]', required=True)
rev_path   = prompt_path('Revenue Input (optional)', required=False)
b2b_path   = prompt_path('B2B Corporate Leads (optional)', required=False)
fy_path    = prompt_path('FY2026 Financial Plans (optional)', required=False)
pa_path    = prompt_path('Plan Approval Sheet (optional)', required=False)

print()
data = {'meta': {'generated': datetime.now().strftime('%Y-%m-%d %H:%M')}}

data['fin23'] = load_sheet_rows(fin23_path, sheet_names=['RAW_DATA', 'RawData'], column_allowlist=FIN23_COLUMNS)
data['rev']   = load_revenue_input_rows(rev_path) if rev_path else []
data['b2b']   = load_sheet_rows(b2b_path) if b2b_path else []
data['fy']    = load_sheet_rows(fy_path) if fy_path else []
data['pa']    = load_sheet_rows(pa_path) if pa_path else []

if not rev_path: print('  Revenue Input skipped.')
if not b2b_path: print('  B2B skipped.')
if not fy_path:  print('  FY2026 skipped.')
if not pa_path:  print('  Plan Approval skipped.')

data['meta']['fin23Rows'] = len(data['fin23'])
data['meta']['revRows']   = len(data['rev'])
data['meta']['b2bRows']   = len(data['b2b'])
data['meta']['fyRows']    = len(data['fy'])
data['meta']['paRows']    = len(data['pa'])

print()
print('Writing data.js...', end=' ', flush=True)
with open('data.js', 'w', encoding='utf-8') as f:
    f.write('window.MARKETING_DATA = ')
    json.dump(data, f, ensure_ascii=False)
    f.write(';')
print('done', flush=True)

print()
print('fin23', len(data['fin23']), 'rev', len(data['rev']), 'b2b', len(data['b2b']),
      'fy', len(data['fy']), 'pa', len(data['pa']))
print(f'Total time: {time.time()-t0:.1f}s')
